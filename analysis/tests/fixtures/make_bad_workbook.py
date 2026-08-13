"""Build a synthetic workbook reproducing two real corruption modes.

Derived from an actual alternative copy of the SRI workbook, which had:
  1. a different state ordering in the anchor sheet, so positional column
     assignment attaches each value to the wrong state; and
  2. firearm mortality values sitting in the column the loader reads as
     median household income, which under positional loading substitutes the
     outcome variable for a predictor.

Both are silent under positional assignment and must raise under keyed joins.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

# Three states, deliberately in different orders between sheets.
_ANCHOR_ORDER = ["Alabama", "Alaska", "Arizona"]
_SHUFFLED_ORDER = ["Arizona", "Alabama", "Alaska"]


def build_bad_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Firearm Morality Rate 2020"
    ws.append(["state", "RATE"])
    for state, rate in zip(_ANCHOR_ORDER, [23.6, 23.5, 16.7]):
        ws.append([state, rate])

    # Same states, different order. Positional assignment misaligns silently;
    # a keyed join must notice.
    ws2 = wb.create_sheet("State Poverty Rates 2020")
    ws2.append(["STATE", "RATE"])
    for state, rate in zip(_SHUFFLED_ORDER, [12.8, 14.9, 9.6]):
        ws2.append([state, rate])

    # A sheet missing one of the anchor states entirely.
    ws3 = wb.create_sheet("Registered Guns")
    ws3.append(["state", "% "])
    ws3.append(["Alabama", 0.0318])
    ws3.append(["Alaska", 0.0214])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    build_bad_workbook(Path(__file__).parent / "bad_workbook.xlsx")
