"""Workbook sheets must be joined on the state key, not by row position."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent / "fixtures"))

from make_bad_workbook import build_bad_workbook

from gun_violence.data import _read_sheet_by_state, _read_sheet_positional


@pytest.fixture
def bad_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "bad_workbook.xlsx"
    build_bad_workbook(path)
    return path


def test_reads_sheet_keyed_by_state_not_row_order(bad_workbook: Path) -> None:
    wb = openpyxl.load_workbook(bad_workbook, data_only=True)
    values = _read_sheet_by_state(wb["State Poverty Rates 2020"], "State Poverty Rates 2020")
    # The sheet lists Arizona first, but the value must follow the state name.
    assert values["Alabama"] == pytest.approx(14.9)
    assert values["Arizona"] == pytest.approx(12.8)


def test_raises_on_duplicate_state_in_sheet(tmp_path: Path) -> None:
    path = tmp_path / "dupes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dupes"
    ws.append(["state", "value"])
    ws.append(["Alabama", 1.0])
    ws.append(["Alabama", 2.0])
    wb.save(path)

    wb2 = openpyxl.load_workbook(path, data_only=True)
    with pytest.raises(ValueError, match="duplicate state"):
        _read_sheet_by_state(wb2["Dupes"], "Dupes")


def test_raises_on_unrecognised_state_name(tmp_path: Path) -> None:
    path = tmp_path / "typo.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Typo"
    ws.append(["state", "value"])
    ws.append(["Alabama", 1.0])
    ws.append(["Alabamaa", 2.0])
    wb.save(path)

    wb2 = openpyxl.load_workbook(path, data_only=True)
    with pytest.raises(ValueError, match="unrecognised state"):
        _read_sheet_by_state(wb2["Typo"], "Typo")


def _sheet_with_values(path: Path, values: list[float]) -> openpyxl.worksheet.worksheet.Worksheet:
    """Build a 50-row keyless sheet (column A is the cached error '#VALUE!')."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Keyless"
    ws.append(["#VALUE!", "value"])
    for v in values:
        ws.append(["#VALUE!", v])
    wb.save(path)
    return openpyxl.load_workbook(path, data_only=True)["Keyless"]


def test_positional_read_accepts_plausible_values(tmp_path: Path) -> None:
    ws = _sheet_with_values(tmp_path / "ok.xlsx", [55000.0] * 50)
    values = _read_sheet_positional(ws, "Keyless", "median_household_income")
    assert len(values) == 50
    assert values[0] == 55000.0


def test_positional_read_rejects_a_column_from_the_wrong_sheet(tmp_path: Path) -> None:
    """The real corruption mode: firearm mortality values in the income column."""
    ws = _sheet_with_values(tmp_path / "wrong.xlsx", [23.6, 23.5, 16.7] + [20.0] * 47)
    with pytest.raises(ValueError, match="outside the plausible range"):
        _read_sheet_positional(ws, "Keyless", "median_household_income")


def test_positional_read_rejects_wrong_row_count(tmp_path: Path) -> None:
    ws = _sheet_with_values(tmp_path / "short.xlsx", [55000.0] * 30)
    with pytest.raises(ValueError, match="expected 50 data rows"):
        _read_sheet_positional(ws, "Keyless", "median_household_income")
