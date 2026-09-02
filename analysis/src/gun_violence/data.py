"""Load, clean, and merge the state-level dataset from raw source files.

Two entry points:

- ``build_dataset(sources)``: build the full 50-state DataFrame from raw inputs
- ``load_dataset(path)``: load a previously-built CSV

The ``build_dataset`` path reads the original SRI Excel workbook, the Mother
Jones Google Sheet CSV, and hard-coded 2020 Census-era population and density
figures.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from .constants import FULL_STATE_NAMES, STATE_ABBR

# ---------------------------------------------------------------------------
# 2020 Census-era reference figures (World Population Review / Census).
# Hard-coded to keep the build reproducible without hitting external services.

POPULATION_2020 = {
    "Alabama": 5_223_121, "Alaska": 738_003, "Arizona": 7_691_212, "Arkansas": 3_133_502,
    "California": 39_345_844, "Colorado": 6_036_620, "Connecticut": 3_702_543,
    "Delaware": 1_069_781, "Florida": 23_659_198, "Georgia": 11_401_288, "Hawaii": 1_430_688,
    "Idaho": 2_058_594, "Illinois": 12_735_249, "Indiana": 7_011_912, "Iowa": 3_246_320,
    "Kansas": 2_989_188, "Kentucky": 4_629_682, "Louisiana": 4_621_500, "Maine": 1_421_310,
    "Maryland": 6_285_380, "Massachusetts": 7_169_608, "Michigan": 10_155_806,
    "Minnesota": 5_863_405, "Mississippi": 2_958_148, "Missouri": 6_297_538,
    "Montana": 1_151_831, "Nebraska": 2_030_421, "Nevada": 3_310_833,
    "New Hampshire": 1_422_166, "New Jersey": 9_590_076, "New Mexico": 2_124_222,
    "New York": 20_003_435, "North Carolina": 11_343_875, "North Dakota": 805_329,
    "Ohio": 11_940_399, "Oklahoma": 4_148_818, "Oregon": 4_281_848, "Pennsylvania": 13_073_016,
    "Rhode Island": 1_118_627, "South Carolina": 5_650_232, "South Dakota": 943_078,
    "Tennessee": 7_378_861, "Texas": 32_101_064, "Utah": 3_574_825, "Vermont": 642_805,
    "Virginia": 8_940_572, "Washington": 8_074_082, "West Virginia": 1_764_892,
    "Wisconsin": 5_988_406, "Wyoming": 590_784,
}

DENSITY_PER_SQ_MI = {
    "Alabama": 103.0, "Alaska": 1.3, "Arizona": 68.0, "Arkansas": 60.0,
    "California": 253.0, "Colorado": 58.0, "Connecticut": 765.0, "Delaware": 549.0,
    "Florida": 441.0, "Georgia": 198.0, "Hawaii": 223.0, "Idaho": 25.0,
    "Illinois": 229.0, "Indiana": 196.0, "Iowa": 58.0, "Kansas": 37.0,
    "Kentucky": 117.0, "Louisiana": 107.0, "Maine": 46.0, "Maryland": 648.0,
    "Massachusetts": 919.0, "Michigan": 180.0, "Minnesota": 74.0, "Mississippi": 63.0,
    "Missouri": 92.0, "Montana": 7.9, "Nebraska": 26.0, "Nevada": 30.0,
    "New Hampshire": 159.0, "New Jersey": 1304.0, "New Mexico": 18.0, "New York": 424.0,
    "North Carolina": 233.0, "North Dakota": 12.0, "Ohio": 292.0, "Oklahoma": 60.0,
    "Oregon": 45.0, "Pennsylvania": 292.0, "Rhode Island": 1082.0,
    "South Carolina": 188.0, "South Dakota": 12.0, "Tennessee": 179.0, "Texas": 123.0,
    "Utah": 44.0, "Vermont": 70.0, "Virginia": 226.0, "Washington": 121.0,
    "West Virginia": 73.0, "Wisconsin": 111.0, "Wyoming": 6.1,
}


@dataclass
class DataSources:
    """Paths to raw source files needed to build the dataset."""

    sri_workbook: Path                 # original SRI Excel workbook
    mother_jones_csv: Path             # Mother Jones mass shootings CSV
    output_csv: Path                   # where to write the merged output
    # CDC firearm components, already committed, so this adds no network
    # dependency to the build. Optional: the build still works without it,
    # simply without the split outcomes.
    components_csv: Path | None = None

    def __post_init__(self) -> None:
        self.sri_workbook = Path(self.sri_workbook)
        self.mother_jones_csv = Path(self.mother_jones_csv)
        self.output_csv = Path(self.output_csv)
        if self.components_csv is not None:
            self.components_csv = Path(self.components_csv)


# ---------------------------------------------------------------------------
# SRI workbook extraction

_SRI_SHEETS_FIRST_COL = {
    # sheet name -> column name in output DataFrame
    "Firearm Morality Rate 2020": "firearm_mortality_rate",
    "Registered Guns": "gun_reg_pct",
    "State Poverty Rates 2020": "poverty_rate",
    "Sucide Rates by State 2020": "suicide_rate",
    "Homicide Rates by State 2020": "homicide_rate",
    "Accident Mortality by State": "accident_mortality_rate",
    "Average Credit Score ": "credit_score",
    "Median House Income v Firearm": "median_household_income",
}


def _read_sheet_by_state(ws, sheet_name: str) -> dict[str, float]:
    """Read a two-column state/value sheet into a dict keyed by state name.

    Replaces the previous positional read (fixed rows 2-51, take column B),
    which assumed every sheet listed the same 50 states in the same order and
    verified nothing. A real alternative copy of this workbook has a different
    state order in one sheet and the outcome variable sitting in the column
    read as median income; positional loading would have merged both silently.
    """
    values: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 2:
            continue
        raw_state, value = row[0], row[1]
        if raw_state is None:
            continue
        state = str(raw_state).strip()
        if not state:
            continue
        if state not in FULL_STATE_NAMES:
            raise ValueError(
                f"{sheet_name}: unrecognised state name {state!r}. "
                "Sheets must key on a full state name."
            )
        if state in values:
            raise ValueError(f"{sheet_name}: duplicate state {state!r}")
        values[state] = value
    return values


# Plausible value range per output column, used to verify sheets whose state
# key is unrecoverable. These are deliberately wide -- the job is to catch a
# whole column coming from the wrong sheet, not to police individual outliers.
# The motivating case: an alternative copy of this workbook had firearm
# mortality values (23.6, 23.5, 16.7) sitting in the median-income column.
_PLAUSIBLE_RANGE = {
    "firearm_mortality_rate": (1.0, 40.0),        # deaths per 100k
    "gun_reg_pct": (0.0, 1.0),                     # a fraction, not a percent
    "poverty_rate": (3.0, 30.0),                   # percent
    "suicide_rate": (3.0, 40.0),                   # per 100k
    "homicide_rate": (0.0, 30.0),                  # per 100k
    "accident_mortality_rate": (15.0, 130.0),      # per 100k
    "credit_score": (500.0, 850.0),                # FICO-like scale
    "median_household_income": (30_000.0, 150_000.0),
}

# Sheets whose column A is the cached Excel error '#VALUE!' rather than a state
# name. Their values are still row-aligned to the anchor sheet, so they are read
# positionally and verified by range instead of by key. Listed explicitly so the
# fallback can never be applied silently to a sheet that ought to have a key.
_KEYLESS_SHEETS = {
    "State Poverty Rates 2020",
    "Sucide Rates by State 2020",
    "Homicide Rates by State 2020",
    "Accident Mortality by State",
    "Median House Income v Firearm",
}


def _read_sheet_positional(ws, sheet_name: str, col_name: str) -> list:
    """Read column B by row position, for sheets with no usable state key.

    Cannot verify alignment -- there is no key to align against. Verifies what
    it can: exactly 50 data rows, and every value inside a plausible range for
    this column. The range check is what catches a column sourced from the
    wrong sheet.

    Row count is measured from the sheet's actual contiguous data block --
    stopping at the first row whose first cell is blank -- rather than via a
    fixed min_row=2/max_row=51 window. openpyxl pads a capped iter_rows()
    range with (None, None) rows past the sheet's real extent, so a window
    read would always report exactly 50 rows regardless of how much real data
    the sheet has, silently defeating this check for short sheets. The real
    workbook's keyless sheets have exactly 50 populated rows followed by
    blank rows and a footer, so this still reads the same 50 values as before.
    """
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            break
        data_rows.append(row)

    if len(data_rows) != 50:
        raise ValueError(f"{sheet_name}: expected 50 data rows, got {len(data_rows)}")

    values = [r[1] if len(r) > 1 else None for r in data_rows]

    lo, hi = _PLAUSIBLE_RANGE[col_name]
    for i, v in enumerate(values):
        if v is None:
            raise ValueError(f"{sheet_name}: empty value at row {i + 2}")
        if not isinstance(v, (int, float)):
            raise TypeError(
                f"{sheet_name}: non-numeric value {v!r} at row {i + 2}"
            )
        if not (lo <= float(v) <= hi):
            raise ValueError(
                f"{sheet_name}: value {v} at row {i + 2} is outside the "
                f"plausible range [{lo}, {hi}] for {col_name}. This usually "
                "means the column came from the wrong sheet."
            )
    return values


def _load_sri_workbook(path: Path) -> pd.DataFrame:
    """Extract state-level columns from the original SRI workbook.

    Sheets that carry a state name in column A are joined on it: a missing,
    duplicated, or unrecognised state raises rather than silently misaligning.

    Five sheets have '#VALUE!' in column A -- broken formulas whose error was
    cached -- so no key survives to join on. Those are read positionally and
    every value is range-checked instead, which catches the corruption mode
    that matters (a whole column sourced from the wrong sheet).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)

    anchor = _read_sheet_by_state(
        wb["Firearm Morality Rate 2020"], "Firearm Morality Rate 2020"
    )
    # Preserve the anchor sheet's own row order: the keyless sheets are aligned
    # to it positionally, so re-sorting here would break them.
    anchor_states = list(anchor)
    df = pd.DataFrame({"state": anchor_states})
    df["firearm_mortality_rate"] = df["state"].map(anchor)

    for sheet_name, col_name in _SRI_SHEETS_FIRST_COL.items():
        if col_name == "firearm_mortality_rate":
            continue  # already taken from the anchor sheet
        ws = wb[sheet_name]
        if sheet_name in _KEYLESS_SHEETS:
            df[col_name] = _read_sheet_positional(ws, sheet_name, col_name)
        else:
            values = _read_sheet_by_state(ws, sheet_name)
            missing = set(df["state"]) - set(values)
            if missing:
                # A keyed sheet may genuinely lack a state. 'Average Credit
                # Score ' has no South Carolina row and instead carries
                # District of Columbia, which this project excludes. Under the
                # previous positional read that inserted DC at row 8 and
                # dropped South Carolina, shifting every state alphabetically
                # between Florida and South Carolina by one row -- 32 states
                # carried another state's credit score. NaN is the correct
                # answer where the source has no value; the alternative is
                # silently attaching a neighbour's.
                print(
                    f"  note: {sheet_name} has no row for {sorted(missing)} "
                    "-- these become NaN"
                )
            df[col_name] = df["state"].map(values)

    # governor party lookup
    ws = wb["us-governors"]
    header = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    state_idx = header.index("state_name")
    party_idx = header.index("party")
    party_map = {row[state_idx]: row[party_idx] for row in rows}
    df["gov_party"] = df["state"].map(party_map)

    return df


# ---------------------------------------------------------------------------
# Mother Jones aggregation

def _parse_state_from_location(loc: str | None) -> str | None:
    """Extract full state name from a 'City, State' string."""
    if loc is None or (isinstance(loc, float) and pd.isna(loc)):
        return None
    tail = str(loc).split(",")[-1].strip().split(";")[0].strip()
    if tail in FULL_STATE_NAMES:
        return tail
    if tail.upper() in STATE_ABBR:
        return STATE_ABBR[tail.upper()]
    return None


def _load_mother_jones(path: Path, year_start: int = 2013) -> pd.DataFrame:
    """Aggregate Mother Jones incidents to state-level counts.

    Uses the 2013-onward window because the definition threshold changed in 2013
    (from 4+ to 3+ fatalities) and mixing the two windows would double-count
    the definition change as a trend.
    """
    df = pd.read_csv(path)
    df.loc[df["location"] == "Baton Rouge, Lousiana", "location"] = "Baton Rouge, Louisiana"
    df.loc[df["location"] == "Washington, D.C.", "location"] = "Washington, DC"
    df["state"] = df["location"].apply(_parse_state_from_location)
    df["year"] = pd.to_numeric(df["year"], errors="coerce")

    recent = df[df["year"] >= year_start]
    unparsed = recent[recent["state"].isna()]
    if len(unparsed) > 0:
        raise ValueError(f"Unparsed locations in Mother Jones data: {unparsed['location'].tolist()}")

    agg = (
        recent.groupby("state")
        .agg(
            mass_shootings_count=("case", "count"),
            mass_shooting_fatalities=("fatalities", "sum"),
        )
        .reset_index()
    )
    return agg[agg["state"] != "District of Columbia"]


# ---------------------------------------------------------------------------
# Public API

def build_dataset(sources: DataSources) -> pd.DataFrame:
    """Build the full merged state-level dataset and write it to CSV.

    Returns the DataFrame; also writes ``sources.output_csv`` as a side effect.
    """
    df = _load_sri_workbook(sources.sri_workbook)

    df["pop_density"] = df["state"].map(DENSITY_PER_SQ_MI)
    df["population"] = df["state"].map(POPULATION_2020)

    ms = _load_mother_jones(sources.mother_jones_csv)
    df = df.merge(ms, on="state", how="left")
    df["mass_shootings_count"] = df["mass_shootings_count"].fillna(0)
    df["mass_shooting_fatalities"] = df["mass_shooting_fatalities"].fillna(0)
    df["mass_shootings_per_10m"] = df["mass_shootings_count"] / (df["population"] / 10_000_000)

    df["gov_party_rep"] = (df["gov_party"] == "republican").astype(int)

    if sources.components_csv is not None and sources.components_csv.exists():
        df = _merge_firearm_components(df, sources.components_csv)

    df = _blank_suppressed_zeros(df)

    _validate(df)
    sources.output_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(sources.output_csv, index=False)
    return df


def load_dataset(path: str | Path) -> pd.DataFrame:
    """Load a previously-built dataset from CSV."""
    df = pd.read_csv(path)
    _validate(df)
    return df


# Columns that must be present AND complete. A gap here means the build is
# broken, not that the underlying figure is unavailable.
REQUIRED_COMPLETE = {
    "state", "firearm_mortality_rate", "gun_reg_pct", "poverty_rate",
    "median_household_income", "pop_density", "population",
    "gov_party_rep", "mass_shootings_count", "mass_shootings_per_10m",
}

# Columns whose source legitimately lacks a value for some state. These may be
# NaN. 'credit_score' is here because the source sheet has no South Carolina
# row at all -- absent must read as absent rather than borrowing a neighbour's
# value, which is precisely the defect that put the wrong score on 32 states.
# CDC mortality rates where an exact 0.0 is a suppressed cell rather than a
# measured zero. NCHS suppresses counts of 1-9, and the workbook recorded those
# suppressed cells as zeros: homicide_rate is exactly 0.0 for New Hampshire and
# Vermont, while the smallest non-zero rate in the column is 1.6. That gap is a
# discontinuity rather than a taper, which is what suppression looks like -- a
# genuinely low rate would have values between. At New Hampshire's 1.4 million
# residents a rate of 1.6 is roughly 23 homicides, so a suppressed cell of
# fewer than 10 puts the true rate somewhere in (0, 0.7]. Zero is the one value
# it cannot be.
#
# This matters because a zero asserts that no event occurred, which is exactly
# the claim the tracker refuses to make when it renders an em dash instead of a
# 0. The same rule now holds on the analysis side.
SUPPRESSIBLE = {
    "homicide_rate", "suicide_rate", "accident_mortality_rate",
    # CDC's firearm components: New Hampshire and Vermont are suppressed in
    # firearm_homicide_rate, the same two states suppressed elsewhere here.
    "firearm_suicide_rate", "firearm_homicide_rate", "firearm_mortality_rate_crude",
}

ALLOWED_MISSING = {"credit_score"} | SUPPRESSIBLE


def _blank_suppressed_zeros(df: pd.DataFrame) -> pd.DataFrame:
    """Turn exact-zero rates in SUPPRESSIBLE columns into NaN.

    Done at load rather than in analysis so no consumer ever sees the zero. A
    suppressed cell is an unknown value, and averaging or regressing on it as
    if it were zero biases the estimate toward zero for precisely the smallest
    states -- the ones most likely to be suppressed in the first place.
    """
    for col in SUPPRESSIBLE & set(df.columns):
        zeros = df[col] == 0.0
        if zeros.any():
            states = df.loc[zeros, "state"].tolist()
            print(f"  note: {col} is exactly 0 for {states} -- recording as "
                  "missing, since a suppressed CDC cell is not a measured zero")
            df.loc[zeros, col] = pd.NA
    return df


def _merge_firearm_components(df: pd.DataFrame, path: Path, year: int = 2020) -> pd.DataFrame:
    """Attach CDC's firearm suicide/homicide split for the cross-section year.

    The workbook's own firearm_mortality_rate is AGE-ADJUSTED while CDC's series
    is CRUDE, so the crude total is carried alongside as
    firearm_mortality_rate_crude rather than replacing it. A component must be
    compared against a total on the same denominator treatment; mixing them
    would confound composition with rate type.

    The workbook's existing suicide_rate and homicide_rate are ALL-CAUSE, not
    firearm-specific -- 16.1 and 7.7 per 100,000 against 9.1 and 5.6 -- so they
    cannot serve as this split.
    """
    comp = pd.read_csv(path)
    comp = comp[comp["year"] == year].drop(columns=["year"])
    comp = comp.rename(columns={"firearm_mortality_rate": "firearm_mortality_rate_crude"})
    merged = df.merge(comp, on="state", how="left")
    missing = merged.loc[merged["firearm_suicide_rate"].isna(), "state"].tolist()
    if missing:
        print(f"  note: no {year} firearm components for {missing}")
    return merged


def _validate(df: pd.DataFrame, *, panel: bool = False) -> None:
    """Sanity checks on the merged dataset.

    With ``panel=False`` the frame is the 50-state cross-section. With
    ``panel=True`` it is state-year, and the shape checks change accordingly:
    one row per state per year, no duplicate (state, year), and every state
    observed in the same years.

    The balanced check is not pedantry. A within-state estimator weights each
    state by how many years it contributes, so an unbalanced panel silently
    changes what the coefficient means -- states with more observations pull
    harder, and which states those are is rarely random.
    """
    if panel:
        if "year" not in df.columns:
            raise ValueError("panel frame has no 'year' column")

        dupes = df.duplicated(subset=["state", "year"])
        if dupes.any():
            examples = (
                df.loc[dupes, ["state", "year"]].head(5).to_dict("records")
            )
            raise ValueError(f"duplicate (state, year) rows: {examples}")

        n_states = df["state"].nunique()
        if n_states != 50:
            raise ValueError(f"Expected 50 states, got {n_states}")

        years_per_state = df.groupby("state")["year"].apply(frozenset)
        distinct = set(years_per_state)
        if len(distinct) != 1:
            counts = df.groupby("state")["year"].nunique()
            short = counts[counts != counts.max()].to_dict()
            raise ValueError(
                "unbalanced panel: not every state is observed in the same "
                f"years. States with fewer: {short}"
            )

        expected = n_states * len(next(iter(distinct)))
        if len(df) != expected:
            raise ValueError(f"Expected {expected} state-years, got {len(df)}")
    elif len(df) != 50:
        raise ValueError(f"Expected 50 states, got {len(df)}")

    missing = (REQUIRED_COMPLETE | ALLOWED_MISSING) - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    nan_cols = [
        c for c in REQUIRED_COMPLETE if c != "state" and df[c].isna().any()
    ]
    if nan_cols:
        raise ValueError(f"NaN values in required columns: {nan_cols}")

    # Belt and braces: _blank_suppressed_zeros should have cleared these at
    # load, so an exact zero surviving to here means a path bypassed it.
    zero_cols = {
        c: df.loc[df[c] == 0.0, "state"].tolist()
        for c in SUPPRESSIBLE & set(df.columns)
        if (df[c] == 0.0).any()
    }
    if zero_cols:
        raise ValueError(
            f"exact zero in suppressible column(s): {zero_cols}. A CDC rate of "
            "exactly 0 is a suppressed cell, not a measured zero; it must be "
            "recorded as missing."
        )
